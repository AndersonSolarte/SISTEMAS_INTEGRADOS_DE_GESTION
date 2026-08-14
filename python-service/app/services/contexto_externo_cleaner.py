from __future__ import annotations

import functools
import io
import os
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from typing import Any, BinaryIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


TABULAR_BASE_HEADERS = [
    "CÓDIGO_INSTITUCIÓN_PADRE",
    "CÓDIGO_INSTITUCIÓN",
    "Institución de Educación Superior (IES)",
    "Principal o Seccional",
    "ID Sector IES",
    "Sector IES",
    "IES Acreditada",
    "ID Caracter",
    "Caracter IES",
    "Código del departamento (IES)",
    "Departamento de domicilio de la IES",
    "Código del Municipio (IES)",
    "Municipio de domicilio de la IES",
    "Código SNIES del programa",
    "Programa Académico",
    "Programa Acreditado",
    "ID Nivel Académico",
    "Nivel Académico",
    "ID Nivel de Formación",
    "Nivel de Formación",
    "ID Metodología",
    "Metodología",
    "ID Área",
    "Área de Conocimiento",
    "Id_Nucleo",
    "Núcleo Básico del Conocimiento (NBC)",
    "ID CINE CAMPO AMPLIO",
    "DESC CINE CAMPO AMPLIO",
    "ID CINE CAMPO ESPECIFICO",
    "DESC CINE CAMPO ESPECIFICO",
    "ID CINE CODIGO DETALLADO",
    "DESC CINE CODIGO DETALLADO",
    "Código del Departamento (Programa)",
    "Departamento de oferta del programa",
    "Código del Municipio (Programa)",
    "Municipio de oferta del programa",
    "ID Sexo",
    "Sexo",
    "Año",
    "Semestre",
]

PROGRAM_HEADERS = [
    "CODIGO_INSTITUCION_PADRE", "CODIGO_INSTITUCION", "NOMBRE_INSTITUCION", "ESTADO_INSTITUCION",
    "CARACTER_ACADEMICO", "SECTOR", "REGISTRO_UNICO", "CODIGO_SNIES_DEL_PROGRAMA",
    "CODIGO_ANTERIOR_ICFES", "NOMBRE_DEL_PROGRAMA", "TITULO_OTORGADO", "ESTADO_PROGRAMA",
    "JUSTIFICACION", "JUSTIFICACION_DETALLADA", "RECONOCIMIENTO_DEL_MINISTERIO",
    "RESOLUCION_DE_APROBACION", "FECHA_DE_RESOLUCION", "FECHA_EJECUTORIA", "VIGENCIA_ANOS",
    "FECHA_DE_REGISTRO_EN_SNIES", "CINE_F_2013_AC_CAMPO_AMPLIO", "CINE_F_2013_AC_CAMPO_ESPECIFIC",
    "CINE_F_2013_AC_CAMPO_DETALLADO", "AREA_DE_CONOCIMIENTO", "NUCLEO_BASICO_DEL_CONOCIMIENTO",
    "NIVEL_ACADEMICO", "NIVEL_DE_FORMACION", "MODALIDAD", "NUMERO_CREDITOS",
    "NUMERO_PERIODOS_DE_DURACION", "PERIODICIDAD", "SE_OFRECE_POR_CICLOS_PROPEDUT",
    "PERIODICIDAD_ADMISIONES", "PROGRAMA_EN_CONVENIO", "DEPARTAMENTO_OFERTA_PROGRAMA",
    "MUNICIPIO_OFERTA_PROGRAMA", "TIPO_CUBRIMIENTO", "COSTO_MATRICULA_ESTUD_NUEVOS", "VIGENCIA_TRANSITORIA",
    "OBSERVACION_DECRETO_1174_23",
]

LIST_CONFIG = {
    "PROGRAMAS_CONTEXTO_EXTERNO": PROGRAM_HEADERS,
    "INSCRITOS_CONTEXTO_EXTERNO": [*TABULAR_BASE_HEADERS, "INSCRITOS"],
    "ADMITIDOS_CONTEXTO_EXTERNO": [*TABULAR_BASE_HEADERS, "ADMITIDOS"],
    "PRIMER_CURSO_CONTEXTO_EXTERNO": [*TABULAR_BASE_HEADERS, "PRIMER CURSO"],
    "MATRICULADOS_CONTEXTO_EXTERNO": [*TABULAR_BASE_HEADERS, "MATRICULADOS"],
    "GRADUADOS_CONTEXTO_EXTERNO": [*TABULAR_BASE_HEADERS, "GRADUADOS"],
}

HEADER_ALIASES = {
    "CÓDIGO_INSTITUCIÓN_PADRE": ["IES PADRE", "CODIGO_INSTITUCION_PADRE", "CÓDIGO_INSTITUCIÓN_PADRE", "CODIGO INSTITUCION PADRE"],
    "CÓDIGO_INSTITUCIÓN": ["CODIGO DE LA INSTITUCION", "CODIGO_INSTITUCION", "CÓDIGO_INSTITUCIÓN", "CODIGO INSTITUCION"],
    "Institución de Educación Superior (IES)": ["INSTITUCION DE EDUCACION SUPERIOR (IES)", "NOMBRE_INSTITUCION", "INSTITUCION", "IES"],
    "Principal o Seccional": ["TIPO IES", "PRINCIPAL O SECCIONAL", "PRINCIPAL_O_SECCIONAL", "PRINCIPAL OSECCIONAL"],
    "ID Sector IES": ["ID SECTOR IES", "ID_SECTOR_IES"],
    "Sector IES": ["SECTOR IES", "SECTOR_IES"],
    "IES Acreditada": ["IES ACREDITADA"],
    "ID Caracter": ["ID CARACTER IES", "ID CARACTER", "ID_CARACTER"],
    "Caracter IES": ["CARACTER IES", "CARACTER_IES"],
    "Código del departamento (IES)": ["CODIGO DEL DEPARTAMENTO (IES)", "CODIGO_DEL_DEPARTAMENTO_(IES)"],
    "Departamento de domicilio de la IES": ["DEPARTAMENTO DE DOMICILIO DE LA IES", "DEPARTAMENTO_DE_DOMICILIO_DE_LA_IES"],
    "Código del Municipio (IES)": ["CODIGO DEL MUNICIPIO IES", "CODIGO_DEL_MUNICIPIO_(IES)", "CODIGO DEL MUNICIPIO (IES)"],
    "Municipio de domicilio de la IES": ["MUNICIPIO DE DOMICILIO DE LA IES", "MUNICIPIO_DE_DOMICILIO_DE_LA_IES"],
    "Código SNIES del programa": ["CODIGO SNIES DEL PROGRAMA", "CODIGO_SNIES_DEL_PROGRAMA", "CODIGO SNIES PROGRAMA"],
    "Programa Académico": ["PROGRAMA ACADEMICO", "NOMBRE_DEL_PROGRAMA", "NOMBRE PROGRAMA", "PROGRAMA"],
    "Programa Acreditado": ["PROGRAMA ACREDITADO"],
    "ID Nivel Académico": ["ID NIVEL ACADEMICO", "ID_NIVEL_ACADEMICO"],
    "Nivel Académico": ["NIVEL ACADEMICO", "NIVEL_ACADEMICO"],
    "ID Nivel de Formación": ["ID NIVEL DE FORMACION", "ID_NIVEL_DE_FORMACION"],
    "Nivel de Formación": ["NIVEL DE FORMACION", "NIVEL_DE_FORMACION"],
    "ID Metodología": ["ID MODALIDAD", "ID METODOLOGIA", "ID_METODOLOGIA"],
    "Metodología": ["MODALIDAD", "METODOLOGIA"],
    "ID Área": ["ID AREA", "ID_AREA"],
    "Área de Conocimiento": ["AREA DE CONOCIMIENTO", "AREA_DE_CONOCIMIENTO"],
    "Id_Nucleo": ["ID NUCLEO", "ID_NUCLEO"],
    "Núcleo Básico del Conocimiento (NBC)": ["NUCLEO BASICO DEL CONOCIMIENTO (NBC)", "NUCLEO_BASICO_DEL_CONOCIMIENTO_(NBC)"],
    "ID CINE CAMPO AMPLIO": ["ID CINE CAMPO AMPLIO"],
    "DESC CINE CAMPO AMPLIO": ["DESC CINE CAMPO AMPLIO"],
    "ID CINE CAMPO ESPECIFICO": ["ID CINE CAMPO ESPECIFICO"],
    "DESC CINE CAMPO ESPECIFICO": ["DESC CINE CAMPO ESPECIFICO"],
    "ID CINE CODIGO DETALLADO": ["ID CINE CAMPO DETALLADO", "ID CINE CODIGO DETALLADO"],
    "DESC CINE CODIGO DETALLADO": ["DESC CINE CAMPO DETALLADO", "DESC CINE CODIGO DETALLADO"],
    "Código del Departamento (Programa)": ["CODIGO DEL DEPARTAMENTO (PROGRAMA)"],
    "Departamento de oferta del programa": ["DEPARTAMENTO DE OFERTA DEL PROGRAMA"],
    "Código del Municipio (Programa)": ["CODIGO DEL MUNICIPIO (PROGRAMA)"],
    "Municipio de oferta del programa": ["MUNICIPIO DE OFERTA DEL PROGRAMA"],
    "ID Sexo": ["ID SEXO", "ID_SEXO"],
    "Sexo": ["SEXO"],
    "Año": ["ANO", "ANIO", "AÑO"],
    "Semestre": ["SEMESTRE", "PERIODO", "PERIODO ACADEMICO"],
}

NUMERIC_HEADERS = {
    "INSCRITOS", "ADMITIDOS", "PRIMER CURSO", "MATRICULADOS", "GRADUADOS",
    "NUMERO_CREDITOS", "NUMERO_PERIODOS_DE_DURACION", "COSTO_MATRICULA_ESTUD_NUEVOS",
}


class CleaningError(ValueError):
    pass


@dataclass
class CleaningResult:
    content: bytes
    input_rows: int
    output_rows: int
    duplicates_removed: int
    empty_rows_removed: int
    source_sheet: str
    matched_columns: int
    corrections_count: int = 0
    corrections: list[dict[str, Any]] | None = None


CANONICAL_RELATIONS = [
    ("CÓDIGO_INSTITUCIÓN_PADRE", "Institución de Educación Superior (IES)", "IES"),
    ("ID Sector IES", "Sector IES", "SECTOR"),
    ("ID Caracter", "Caracter IES", "CARACTER_IES"),
    ("Código del departamento (IES)", "Departamento de domicilio de la IES", "DEPARTAMENTO"),
    ("Código del Municipio (IES)", "Municipio de domicilio de la IES", "MUNICIPIO"),
    ("Código SNIES del programa", "Programa Académico", "PROGRAMA"),
    ("ID Nivel Académico", "Nivel Académico", "NIVEL_ACADEMICO"),
    ("ID Nivel de Formación", "Nivel de Formación", "NIVEL_FORMACION"),
    ("ID Metodología", "Metodología", "MODALIDAD"),
    ("ID Área", "Área de Conocimiento", "AREA_CONOCIMIENTO"),
    ("Id_Nucleo", "Núcleo Básico del Conocimiento (NBC)", "NUCLEO_BASICO"),
    ("ID CINE CAMPO AMPLIO", "DESC CINE CAMPO AMPLIO", "CINE_AMPLIO"),
    ("ID CINE CAMPO ESPECIFICO", "DESC CINE CAMPO ESPECIFICO", "CINE_ESPECIFICO"),
    ("ID CINE CODIGO DETALLADO", "DESC CINE CODIGO DETALLADO", "CINE_DETALLADO"),
    ("Código del Departamento (Programa)", "Departamento de oferta del programa", "DEPARTAMENTO"),
    ("Código del Municipio (Programa)", "Municipio de oferta del programa", "MUNICIPIO"),
    ("ID Sexo", "Sexo", "SEXO"),
]

GLOBAL_CANONICAL_HEADERS = {value_header for _, value_header, _ in CANONICAL_RELATIONS}

OFFICIAL_METADATA_MAPS = {
    ("ID Sector IES", "Sector IES"): {
        "1": "OFICIAL", "01": "OFICIAL",
        "2": "PRIVADO", "02": "PRIVADO",
    },
    ("ID Caracter", "Caracter IES"): {
        "1": "INSTITUCIÓN TÉCNICA PROFESIONAL", "01": "INSTITUCIÓN TÉCNICA PROFESIONAL",
        "2": "INSTITUCIÓN TECNOLÓGICA", "02": "INSTITUCIÓN TECNOLÓGICA",
        "3": "INSTITUCIÓN UNIVERSITARIA / ESCUELA TECNOLÓGICA", "03": "INSTITUCIÓN UNIVERSITARIA / ESCUELA TECNOLÓGICA",
        "4": "UNIVERSIDAD", "04": "UNIVERSIDAD",
    },
    ("ID Nivel Académico", "Nivel Académico"): {
        "0": "SIN INFORMACIÓN", "00": "SIN INFORMACIÓN",
        "1": "PREGRADO", "01": "PREGRADO",
        "2": "POSGRADO", "02": "POSGRADO",
    },
    ("ID Nivel de Formación", "Nivel de Formación"): {
        "0": "SIN INFORMACIÓN", "00": "SIN INFORMACIÓN",
        "1": "ESPECIALIZACIÓN UNIVERSITARIA", "01": "ESPECIALIZACIÓN UNIVERSITARIA",
        "2": "MAESTRÍA", "02": "MAESTRÍA",
        "3": "DOCTORADO", "03": "DOCTORADO",
        "4": "FORMACIÓN TÉCNICA PROFESIONAL", "04": "FORMACIÓN TÉCNICA PROFESIONAL",
        "5": "TECNOLÓGICO", "05": "TECNOLÓGICO",
        "6": "UNIVERSITARIO", "06": "UNIVERSITARIO",
        "7": "ESPECIALIZACIÓN TÉCNICO PROFESIONAL", "07": "ESPECIALIZACIÓN TÉCNICO PROFESIONAL",
        "8": "ESPECIALIZACIÓN TECNOLÓGICA", "08": "ESPECIALIZACIÓN TECNOLÓGICA",
        "10": "ESPECIALIZACIÓN MÉDICO QUIRÚRGICA",
    },
    ("ID Metodología", "Metodología"): {
        "0": "SIN INFORMACIÓN", "00": "SIN INFORMACIÓN",
        "1": "PRESENCIAL", "01": "PRESENCIAL",
        "2": "A DISTANCIA", "02": "A DISTANCIA",
        "3": "VIRTUAL", "03": "VIRTUAL",
        "4": "PRESENCIAL-VIRTUAL", "04": "PRESENCIAL-VIRTUAL",
        "5": "DUAL", "05": "DUAL",
        "6": "PRESENCIAL-A DISTANCIA", "06": "PRESENCIAL-A DISTANCIA",
        "7": "PRESENCIAL-DUAL", "07": "PRESENCIAL-DUAL",
        "8": "VIRTUAL - A DISTANCIA", "08": "VIRTUAL - A DISTANCIA",
        "9": "VIRTUAL-DUAL", "09": "VIRTUAL-DUAL",
        "11": "PRESENCIAL-VIRTUAL-A DISTANCIA",
        "16": "HÍBRIDA (PRESENCIAL-VIRTUAL)",
        "17": "HÍBRIDA (A DISTANCIA-VIRTUAL)",
    },
    ("ID Sexo", "Sexo"): {
        "1": "MASCULINO", "01": "MASCULINO",
        "2": "FEMENINO", "02": "FEMENINO",
        "3": "NO BINARIO", "03": "NO BINARIO",
        "4": "TRANS", "04": "TRANS",
    },
    ("ID Área", "Área de Conocimiento"): {
        "1": "AGRONOMÍA, VETERINARIA Y AFINES", "01": "AGRONOMÍA, VETERINARIA Y AFINES",
        "2": "BELLAS ARTES", "02": "BELLAS ARTES",
        "3": "CIENCIAS DE LA EDUCACIÓN", "03": "CIENCIAS DE LA EDUCACIÓN",
        "4": "CIENCIAS DE LA SALUD", "04": "CIENCIAS DE LA SALUD",
        "5": "CIENCIAS SOCIALES Y HUMANAS", "05": "CIENCIAS SOCIALES Y HUMANAS",
        "6": "ECONOMÍA, ADMINISTRACIÓN, CONTADURÍA Y AFINES", "06": "ECONOMÍA, ADMINISTRACIÓN, CONTADURÍA Y AFINES",
        "8": "INGENIERÍA, ARQUITECTURA, URBANISMO Y AFINES", "08": "INGENIERÍA, ARQUITECTURA, URBANISMO Y AFINES",
        "9": "MATEMÁTICAS Y CIENCIAS NATURALES", "09": "MATEMÁTICAS Y CIENCIAS NATURALES",
    },
}


def normalize_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.upper().replace("Ñ", "N")
    return re.sub(r"[^A-Z0-9]+", "_", text).strip("_")


def fix_corrupted_encoding(value: Any) -> Any:
    if not isinstance(value, str) or not value:
        return value
    t = value
    BAD = r'[\xbf\ufffd¿\?]'

    # 1. Words with DISEÑO / DISEÑADOR / SEÑAL / Ñ / GRÁFICO
    t = re.sub(fr'DISE{BAD}O', 'DISEÑO', t, flags=re.IGNORECASE)
    t = re.sub(fr'DIS{BAD}O', 'DISEÑO', t, flags=re.IGNORECASE)
    t = re.sub(fr'DISE{BAD}AD', 'DISEÑAD', t, flags=re.IGNORECASE)
    t = re.sub(fr'DIS{BAD}AD', 'DISEÑAD', t, flags=re.IGNORECASE)
    t = re.sub(fr'DISE{BAD}A', 'DISEÑA', t, flags=re.IGNORECASE)
    t = re.sub(fr'GR{BAD}FIC', 'GRÁFIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'NI{BAD}EZ', 'NIÑEZ', t, flags=re.IGNORECASE)
    t = re.sub(fr'NI{BAD}O', 'NIÑO', t, flags=re.IGNORECASE)
    t = re.sub(fr'NI{BAD}A', 'NIÑA', t, flags=re.IGNORECASE)
    t = re.sub(fr'ESPA{BAD}OL', 'ESPAÑOL', t, flags=re.IGNORECASE)
    t = re.sub(fr'\bA{BAD}O', 'AÑO', t, flags=re.IGNORECASE)
    t = re.sub(fr'\bA{BAD}OS', 'AÑOS', t, flags=re.IGNORECASE)
    t = re.sub(fr'NARI{BAD}O', 'NARIÑO', t, flags=re.IGNORECASE)

    # 2. Words ending in IÓN / CIÓN / SIÓN / TIÓN
    t = re.sub(fr'I{BAD}N\b', 'IÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'C{BAD}N\b', 'CIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'CI{BAD}N', 'CIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'SI{BAD}N', 'SIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'TI{BAD}N', 'TIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'GESTI{BAD}N', 'GESTIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'DIRECCI{BAD}N', 'DIRECCIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'PLANEACI{BAD}N', 'PLANEACIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'PRODUCCI{BAD}N', 'PRODUCCIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'CONSTRUCCI{BAD}N', 'CONSTRUCCIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'EVALUACI{BAD}N', 'EVALUACIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'ACREDITACI{BAD}N', 'ACREDITACIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'INVESTIGACI{BAD}N', 'INVESTIGACIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'ESPECIALIZACI{BAD}N', 'ESPECIALIZACIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'FUNDACI{BAD}N', 'FUNDACIÓN', t, flags=re.IGNORECASE)
    t = re.sub(fr'INSTITUCI{BAD}N', 'INSTITUCIÓN', t, flags=re.IGNORECASE)

    # 3. Words ending in ÍA
    t = re.sub(fr'([B-DF-HJ-NP-TV-Z]){BAD}A\b', r'\1ÍA', t, flags=re.IGNORECASE)
    t = re.sub(fr'BACTERIOL{BAD}G{BAD}A', 'BACTERIOLOGÍA', t, flags=re.IGNORECASE)
    t = re.sub(fr'BACTERIOLOG{BAD}A', 'BACTERIOLOGÍA', t, flags=re.IGNORECASE)
    t = re.sub(fr'BIOL{BAD}G{BAD}A', 'BIOLOGÍA', t, flags=re.IGNORECASE)
    t = re.sub(fr'BIOLOG{BAD}A', 'BIOLOGÍA', t, flags=re.IGNORECASE)
    t = re.sub(fr'LOG{BAD}A\b', 'LOGÍA', t, flags=re.IGNORECASE)
    t = re.sub(fr'INGENIER{BAD}A', 'INGENIERÍA', t, flags=re.IGNORECASE)
    t = re.sub(fr'ENFERMER{BAD}A', 'ENFERMERÍA', t, flags=re.IGNORECASE)
    t = re.sub(fr'MAESTR{BAD}A', 'MAESTRÍA', t, flags=re.IGNORECASE)
    t = re.sub(fr'OPTOMETR{BAD}A', 'OPTOMETRÍA', t, flags=re.IGNORECASE)
    t = re.sub(fr'TECNOLOG{BAD}A', 'TECNOLOGÍA', t, flags=re.IGNORECASE)

    # 4. Specific Colombian academic terms with corrupted accents
    t = re.sub(fr'AN{BAD}LISIS', 'ANÁLISIS', t, flags=re.IGNORECASE)
    t = re.sub(fr'MEC{BAD}NIC', 'MECÁNIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'TELEM{BAD}TIC', 'TELEMÁTIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'M{BAD}DIC', 'MÉDIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'ESTRAT{BAD}GIC', 'ESTRATÉGIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'DEMOGR{BAD}FIC', 'DEMOGRÁFIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'ECON{BAD}MIC', 'ECONÓMIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'POL{BAD}TIC', 'POLÍTICO', t, flags=re.IGNORECASE)
    t = re.sub(fr'QUIR{BAD}RGIC', 'QUIRÚRGIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'PROPED{BAD}UTIC', 'PROPEDÉUTIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'{BAD}NFASIS', 'ÉNFASIS', t, flags=re.IGNORECASE)
    t = re.sub(fr'INGL{BAD}S', 'INGLÉS', t, flags=re.IGNORECASE)
    t = re.sub(fr'FRANC{BAD}S', 'FRANCÉS', t, flags=re.IGNORECASE)
    t = re.sub(fr'F{BAD}SIC', 'FÍSIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'P{BAD}BLIC', 'PÚBLIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'QU{BAD}MIC', 'QUÍMIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'LING[ÜÜ{BAD}]+STIC', 'LINGÜÍSTIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'TR{BAD}PIC', 'TRÓPIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'T{BAD}CNIC', 'TÉCNIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'TECNOL{BAD}GIC', 'TECNOLÓGIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'AGRON{BAD}MIC', 'AGRONÓMIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'AGUST{BAD}N', 'AGUSTÍN', t, flags=re.IGNORECASE)
    t = re.sub(fr'TUR{BAD}STIC', 'TURÍSTIC', t, flags=re.IGNORECASE)
    t = re.sub(fr'BILING{BAD}E', 'BILINGÜE', t, flags=re.IGNORECASE)

    # 5. Clean standalone separators like ¿ MBA ¿ or ¿ PERIODISMO
    t = re.sub(fr'\bMBA\s*{BAD}', 'MBA', t, flags=re.IGNORECASE)
    t = re.sub(fr'\s*{BAD}\s*', ' - ', t)
    t = re.sub(fr'([A-Za-záéíóúÁÉÍÓÚñÑ]){BAD}([A-Za-záéíóúÁÉÍÓÚñÑ])', r'\1I\2', t)
    return t.strip(' -')


# ---------------------------------------------------------------------------
# fix_missing_accents: restore tildes to words typed without accents
# ---------------------------------------------------------------------------
# Maps: unaccented regex pattern → accented replacement
# All patterns use word boundaries (\b) to avoid partial matches.
_ACCENT_RULES: list[tuple[str, str, int]] = [
    # ── Terminaciones en -CION / -SION (sin tilde) → -CIÓN / -SIÓN ──
    (r'\bESPECIALIZACION\b', 'ESPECIALIZACIÓN', 0),
    (r'\bADMINISTRACION\b', 'ADMINISTRACIÓN', 0),
    (r'\bCOMUNICACION\b', 'COMUNICACIÓN', 0),
    (r'\bEDUCACION\b', 'EDUCACIÓN', 0),
    (r'\bINVESTIGACION\b', 'INVESTIGACIÓN', 0),
    (r'\bINSTITUCION\b', 'INSTITUCIÓN', 0),
    (r'\bFUNDACION\b', 'FUNDACIÓN', 0),
    (r'\bCORPORACION\b', 'CORPORACIÓN', 0),
    (r'\bCONSTRUCCION\b', 'CONSTRUCCIÓN', 0),
    (r'\bPRODUCCION\b', 'PRODUCCIÓN', 0),
    (r'\bDIRECCION\b', 'DIRECCIÓN', 0),
    (r'\bPLANEACION\b', 'PLANEACIÓN', 0),
    (r'\bEVALUACION\b', 'EVALUACIÓN', 0),
    (r'\bACREDITACION\b', 'ACREDITACIÓN', 0),
    (r'\bGESTION\b', 'GESTIÓN', 0),
    (r'\bPREVENCION\b', 'PREVENCIÓN', 0),
    (r'\bATENCION\b', 'ATENCIÓN', 0),
    (r'\bNUTRICION\b', 'NUTRICIÓN', 0),
    (r'\bANIMACION\b', 'ANIMACIÓN', 0),
    (r'\bREHABILITACION\b', 'REHABILITACIÓN', 0),
    (r'\bFORMACION\b', 'FORMACIÓN', 0),
    (r'\bCONTAMINACION\b', 'CONTAMINACIÓN', 0),
    (r'\bRELACION\b', 'RELACIÓN', 0),
    (r'\bINFORMACION\b', 'INFORMACIÓN', 0),
    (r'\bORGANIZACION\b', 'ORGANIZACIÓN', 0),
    (r'\bOPERACION\b', 'OPERACIÓN', 0),
    (r'\bNAVEGACION\b', 'NAVEGACIÓN', 0),
    (r'\bAVIACION\b', 'AVIACIÓN', 0),
    (r'\bCIVILIZACION\b', 'CIVILIZACIÓN', 0),
    (r'\bLEGISLACION\b', 'LEGISLACIÓN', 0),
    (r'\bREGULACION\b', 'REGULACIÓN', 0),
    (r'\bOCUPACION\b', 'OCUPACIÓN', 0),
    (r'\bINNOVACION\b', 'INNOVACIÓN', 0),
    (r'\bAUTOMATIZACION\b', 'AUTOMATIZACIÓN', 0),
    (r'\bTELECOMUNICACION\b', 'TELECOMUNICACIÓN', 0),
    (r'\bTRANSFORMACION\b', 'TRANSFORMACIÓN', 0),
    (r'\bNEGOCIACION\b', 'NEGOCIACIÓN', 0),
    (r'\bINTERVENCION\b', 'INTERVENCIÓN', 0),
    (r'\bPROTECCION\b', 'PROTECCIÓN', 0),
    (r'\bALIMENTACION\b', 'ALIMENTACIÓN', 0),
    (r'\bCOOPERACION\b', 'COOPERACIÓN', 0),
    (r'\bINTEGRACION\b', 'INTEGRACIÓN', 0),
    (r'\bPROFESION\b', 'PROFESIÓN', 0),
    (r'\bEXTENSION\b', 'EXTENSIÓN', 0),
    (r'\bPROGRAMACION\b', 'PROGRAMACIÓN', 0),
    (r'\bRECREACION\b', 'RECREACIÓN', 0),
    (r'\bPOBLACION\b', 'POBLACIÓN', 0),
    (r'\bINSTALACION\b', 'INSTALACIÓN', 0),
    (r'\bCONSERVACION\b', 'CONSERVACIÓN', 0),
    (r'\bEXPLOTACION\b', 'EXPLOTACIÓN', 0),
    (r'\bTRIBUTACION\b', 'TRIBUTACIÓN', 0),
    (r'\bCREACION\b', 'CREACIÓN', 0),
    (r'\bNEGACION\b', 'NEGACIÓN', 0),
    (r'\bORIENTACION\b', 'ORIENTACIÓN', 0),
    (r'\bELABORACION\b', 'ELABORACIÓN', 0),
    (r'\bCONFIGURACION\b', 'CONFIGURACIÓN', 0),
    (r'\bSISTEMATIZACION\b', 'SISTEMATIZACIÓN', 0),
    (r'\bMOTIVACION\b', 'MOTIVACIÓN', 0),
    (r'\bMEDIACION\b', 'MEDIACIÓN', 0),
    (r'\bCONCILIACION\b', 'CONCILIACIÓN', 0),
    (r'\bNEGOCIACION\b', 'NEGOCIACIÓN', 0),
    (r'\bCONTRATACION\b', 'CONTRATACIÓN', 0),
    (r'\bLIQUIDACION\b', 'LIQUIDACIÓN', 0),
    (r'\bINSPECCION\b', 'INSPECCIÓN', 0),
    (r'\bAPLICACION\b', 'APLICACIÓN', 0),
    (r'\bSECCIONAL\b', 'SECCIONAL', 0),
    # Generic -CION / -SION catch-all (after specific words)
    (r'(?<=[A-Z])ACION\b', 'ACIÓN', 0),
    (r'(?<=[A-Z])ECION\b', 'ECCIÓN', 0),
    (r'(?<=[A-Z])ICION\b', 'ICIÓN', 0),
    (r'(?<=[A-Z])OCION\b', 'OCIÓN', 0),
    (r'(?<=[A-Z])UCION\b', 'UCIÓN', 0),
    (r'(?<=[A-Z])ASION\b', 'ASIÓN', 0),
    (r'(?<=[A-Z])ESION\b', 'ESIÓN', 0),
    (r'(?<=[A-Z])ISION\b', 'ISIÓN', 0),
    (r'(?<=[A-Z])OSION\b', 'OSIÓN', 0),
    (r'(?<=[A-Z])USION\b', 'USIÓN', 0),
    (r'(?<=[A-Z])SION\b', 'SIÓN', 0),
    # ── Palabras con Ñ ──
    (r'\bDISENO\b', 'DISEÑO', 0),
    (r'\bNARINO\b', 'NARIÑO', 0),
    (r'\bENSENANZA\b', 'ENSEÑANZA', 0),
    (r'\bPEQUENO\b', 'PEQUEÑO', 0),
    (r'\bPEQUENA\b', 'PEQUEÑA', 0),
    (r'\bESPANOL\b', 'ESPAÑOL', 0),
    (r'\bESPANOLA\b', 'ESPAÑOLA', 0),
    (r'\bCANON\b', 'CAÑÓN', 0),
    (r'\bSENAL\b', 'SEÑAL', 0),
    (r'\bMONTANA\b', 'MONTAÑA', 0),
    (r'\bCOMPANIA\b', 'COMPAÑÍA', 0),
    (r'\bINGENIERIA\b', 'INGENIERÍA', 0),
    # ── Palabras con tilde en -ÍA ──
    (r'\bTECNOLOGIA\b', 'TECNOLOGÍA', 0),
    (r'\bBIOLOGIA\b', 'BIOLOGÍA', 0),
    (r'\bSOCIOLOGIA\b', 'SOCIOLOGÍA', 0),
    (r'\bPSICOLOGIA\b', 'PSICOLOGÍA', 0),
    (r'\bFILOSOFIA\b', 'FILOSOFÍA', 0),
    (r'\bTEOLOGIA\b', 'TEOLOGÍA', 0),
    (r'\bECOLOGIA\b', 'ECOLOGÍA', 0),
    (r'\bFISIOLOGIA\b', 'FISIOLOGÍA', 0),
    (r'\bODONTOLOGIA\b', 'ODONTOLOGÍA', 0),
    (r'\bFARMACOLOGIA\b', 'FARMACOLOGÍA', 0),
    (r'\bCRIMINOLOGIA\b', 'CRIMINOLOGÍA', 0),
    (r'\bEPIDEMIOLOGIA\b', 'EPIDEMIOLOGÍA', 0),
    (r'\bGEOLOGIA\b', 'GEOLOGÍA', 0),
    (r'\bZOOLOGIA\b', 'ZOOLOGÍA', 0),
    (r'\bMETEOROLOGIA\b', 'METEOROLOGÍA', 0),
    (r'\bBACTERIOLOGIA\b', 'BACTERIOLOGÍA', 0),
    (r'\bFONOAUDIOLOGIA\b', 'FONOAUDIOLOGÍA', 0),
    (r'\bENFERMERIA\b', 'ENFERMERÍA', 0),
    (r'\bMAESTRIA\b', 'MAESTRÍA', 0),
    (r'\bOPTOMETRIA\b', 'OPTOMETRÍA', 0),
    (r'\bAGRONOMIA\b', 'AGRONOMÍA', 0),
    (r'\bASTRONOMIA\b', 'ASTRONOMÍA', 0),
    (r'\bECONOMIA\b', 'ECONOMÍA', 0),
    (r'\bANATOMIA\b', 'ANATOMÍA', 0),
    (r'\bGASTRONOMIA\b', 'GASTRONOMÍA', 0),
    (r'\bFISIOTERAPIA\b', 'FISIOTERAPIA', 0),
    (r'\bPEDAGOGIA\b', 'PEDAGOGÍA', 0),
    (r'\bGEOGRAFIA\b', 'GEOGRAFÍA', 0),
    (r'\bFOTOGRAFIA\b', 'FOTOGRAFÍA', 0),
    (r'\bTOPOGRAFIA\b', 'TOPOGRAFÍA', 0),
    (r'\bCARTOGRAFIA\b', 'CARTOGRAFÍA', 0),
    (r'\bHIDRAULICA\b', 'HIDRÁULICA', 0),
    (r'\bHIDROLOGIA\b', 'HIDROLOGÍA', 0),
    (r'\bCIRUGIA\b', 'CIRUGÍA', 0),
    (r'\bAUDITORIA\b', 'AUDITORÍA', 0),
    (r'\bCONTADURIA\b', 'CONTADURÍA', 0),
    (r'\bSABIDURIA\b', 'SABIDURÍA', 0),
    (r'\bGANADERIA\b', 'GANADERÍA', 0),
    (r'\bPANADERIA\b', 'PANADERÍA', 0),
    # Generic -LOGIA catch-all
    (r'(?<=[A-Z])LOGIA\b', 'LOGÍA', 0),
    (r'(?<=[A-Z])GRAFIA\b', 'GRAFÍA', 0),
    # ── Palabras con tilde en É ──
    (r'\bELECTRONICA\b', 'ELECTRÓNICA', 0),
    (r'\bMECANICA\b', 'MECÁNICA', 0),
    (r'\bSISTEMATICA\b', 'SISTEMÁTICA', 0),
    (r'\bMATEMATICA\b', 'MATEMÁTICA', 0),
    (r'\bMATEMATICAS\b', 'MATEMÁTICAS', 0),
    (r'\bINFORMATICA\b', 'INFORMÁTICA', 0),
    (r'\bTELEMATICA\b', 'TELEMÁTICA', 0),
    (r'\bESTADISTICA\b', 'ESTADÍSTICA', 0),
    (r'\bLOGISTICA\b', 'LOGÍSTICA', 0),
    (r'\bLINGUISTICA\b', 'LINGÜÍSTICA', 0),
    (r'\bGENETICA\b', 'GENÉTICA', 0),
    (r'\bDIDAC?TICA\b', 'DIDÁCTICA', 0),
    (r'\bPRACTICA\b', 'PRÁCTICA', 0),
    (r'\bPRACTICAS\b', 'PRÁCTICAS', 0),
    (r'\bPUBLICA\b', 'PÚBLICA', 0),
    (r'\bPUBLICO\b', 'PÚBLICO', 0),
    (r'\bACADEMICO\b', 'ACADÉMICO', 0),
    (r'\bACADEMICA\b', 'ACADÉMICA', 0),
    (r'\bECONOMICO\b', 'ECONÓMICO', 0),
    (r'\bECONOMICA\b', 'ECONÓMICA', 0),
    (r'\bECONOMICAS\b', 'ECONÓMICAS', 0),
    (r'\bTECNICO\b', 'TÉCNICO', 0),
    (r'\bTECNICA\b', 'TÉCNICA', 0),
    (r'\bTECNOLOGICO\b', 'TECNOLÓGICO', 0),
    (r'\bTECNOLOGICA\b', 'TECNOLÓGICA', 0),
    (r'\bTECNOLOGICAS\b', 'TECNOLÓGICAS', 0),
    (r'\bPEDAGOGICO\b', 'PEDAGÓGICO', 0),
    (r'\bPEDAGOGICA\b', 'PEDAGÓGICA', 0),
    (r'\bAGRONOMICO\b', 'AGRONÓMICO', 0),
    (r'\bAGRONOMICA\b', 'AGRONÓMICA', 0),
    (r'\bANALISIS\b', 'ANÁLISIS', 0),
    (r'\bSINTESIS\b', 'SÍNTESIS', 0),
    (r'\bGRAFICO\b', 'GRÁFICO', 0),
    (r'\bGRAFICA\b', 'GRÁFICA', 0),
    (r'\bMUSICA\b', 'MÚSICA', 0),
    (r'\bFISICA\b', 'FÍSICA', 0),
    (r'\bQUIMICA\b', 'QUÍMICA', 0),
    (r'\bQUIRURGICA\b', 'QUIRÚRGICA', 0),
    (r'\bBASICA\b', 'BÁSICA', 0),
    (r'\bBASICO\b', 'BÁSICO', 0),
    (r'\bCLASICA\b', 'CLÁSICA', 0),
    (r'\bDEMOGRAFICA\b', 'DEMOGRÁFICA', 0),
    (r'\bGEOGRAFICA\b', 'GEOGRÁFICA', 0),
    (r'\bCATOLICA\b', 'CATÓLICA', 0),
    (r'\bCATOLICO\b', 'CATÓLICO', 0),
    (r'\bAUTONOMA\b', 'AUTÓNOMA', 0),
    (r'\bAUTONOMO\b', 'AUTÓNOMO', 0),
    # ── Adjetivos derivados de -LOGÍA / -GRAFÍA ──
    (r'\bCRIMINOLOGICAS\b', 'CRIMINOLÓGICAS', 0),
    (r'\bCRIMINOLOGICO\b', 'CRIMINOLÓGICO', 0),
    (r'\bCRIMINOLOGICA\b', 'CRIMINOLÓGICA', 0),
    (r'\bMETODOLOGICO\b', 'METODOLÓGICO', 0),
    (r'\bMETODOLOGICA\b', 'METODOLÓGICA', 0),
    (r'\bDEONTOLOGICO\b', 'DEONTOLÓGICO', 0),
    (r'\bDEONTOLOGICA\b', 'DEONTOLÓGICA', 0),
    (r'\bBIOLOGICO\b', 'BIOLÓGICO', 0),
    (r'\bBIOLOGICA\b', 'BIOLÓGICA', 0),
    (r'\bBIOLOGICAS\b', 'BIOLÓGICAS', 0),
    (r'\bGEOGRAFICO\b', 'GEOGRÁFICO', 0),
    (r'\bGEOGRAFICA\b', 'GEOGRÁFICA', 0),
    (r'\bDEMOGRAFICO\b', 'DEMOGRÁFICO', 0),
    (r'\bSOCIOLOGICO\b', 'SOCIOLÓGICO', 0),
    (r'\bSOCIOLOGICA\b', 'SOCIOLÓGICA', 0),
    (r'\bPSICOLOGICO\b', 'PSICOLÓGICO', 0),
    (r'\bPSICOLOGICA\b', 'PSICOLÓGICA', 0),
    (r'\bFILOLOGICO\b', 'FILOLÓGICO', 0),
    (r'\bFILOLOGICA\b', 'FILOLÓGICA', 0),
    (r'\bFARMACOLOGICO\b', 'FARMACOLÓGICO', 0),
    (r'\bFARMACOLOGICA\b', 'FARMACOLÓGICA', 0),
    (r'\bONCOLOGICO\b', 'ONCOLÓGICO', 0),
    (r'\bONCOLOGICA\b', 'ONCOLÓGICA', 0),
    # Generic -LOGICO/-LOGICA/-LOGICAS catch-all
    (r'(?<=[A-Z])LOGICO\b', 'LÓGICO', 0),
    (r'(?<=[A-Z])LOGICA\b', 'LÓGICA', 0),
    (r'(?<=[A-Z])LOGICAS\b', 'LÓGICAS', 0),
    (r'(?<=[A-Z])LOGICOS\b', 'LÓGICOS', 0),
    # ── Más adjetivos con tilde ──
    (r'\bJURIDICO\b', 'JURÍDICO', 0),
    (r'\bJURIDICA\b', 'JURÍDICA', 0),
    (r'\bJURIDICAS\b', 'JURÍDICAS', 0),
    (r'\bMEDICO\b', 'MÉDICO', 0),
    (r'\bMEDICA\b', 'MÉDICA', 0),
    (r'\bMEDICAS\b', 'MÉDICAS', 0),
    (r'\bELECTRONICO\b', 'ELECTRÓNICO', 0),
    (r'\bESTRATEGICO\b', 'ESTRATÉGICO', 0),
    (r'\bESTRATEGICA\b', 'ESTRATÉGICA', 0),
    (r'\bPOLITICO\b', 'POLÍTICO', 0),
    (r'\bPOLITICA\b', 'POLÍTICA', 0),
    (r'\bPOLITICAS\b', 'POLÍTICAS', 0),
    (r'\bCLINICO\b', 'CLÍNICO', 0),
    (r'\bCLINICA\b', 'CLÍNICA', 0),
    (r'\bSISTEMICO\b', 'SISTÉMICO', 0),
    (r'\bSISTEMICA\b', 'SISTÉMICA', 0),
    (r'\bDINAMICO\b', 'DINÁMICO', 0),
    (r'\bDINAMICA\b', 'DINÁMICA', 0),
    (r'\bHISPANICO\b', 'HISPÁNICO', 0),
    (r'\bHISPANICA\b', 'HISPÁNICA', 0),
    (r'\bORGANICO\b', 'ORGÁNICO', 0),
    (r'\bORGANICA\b', 'ORGÁNICA', 0),
    (r'\bBOTANICO\b', 'BOTÁNICO', 0),
    (r'\bBOTANICA\b', 'BOTÁNICA', 0),
    (r'\bAERONAUTICO\b', 'AERONÁUTICO', 0),
    (r'\bAERONAUTICA\b', 'AERONÁUTICA', 0),
    (r'\bNAUTICO\b', 'NÁUTICO', 0),
    (r'\bNAUTICA\b', 'NÁUTICA', 0),
    (r'\bAGRICOLA\b', 'AGRÍCOLA', 0),
    (r'\bPECUARIO\b', 'PECUARIO', 0),
    (r'\bAMBIENTAL\b', 'AMBIENTAL', 0),
    # ── Ciudades y departamentos colombianos ──
    (r'\bBOGOTA\b', 'BOGOTÁ', 0),
    (r'\bBOYACA\b', 'BOYACÁ', 0),
    (r'\bCHOCO\b', 'CHOCÓ', 0),
    (r'\bCORDOBA\b', 'CÓRDOBA', 0),
    (r'\bBOLIVAR\b', 'BOLÍVAR', 0),
    (r'\bCUCUTA\b', 'CÚCUTA', 0),
    (r'\bQUINDIO\b', 'QUINDÍO', 0),
    (r'\bGUAINIA\b', 'GUAINÍA', 0),
    (r'\bVAUPES\b', 'VAUPÉS', 0),
    (r'\bATLANTICO\b', 'ATLÁNTICO', 0),
    (r'\bPACIFICO\b', 'PACÍFICO', 0),
]

# Pre-compile all accent rules for performance
_COMPILED_ACCENT_RULES = [(re.compile(pattern, re.IGNORECASE), replacement) for pattern, replacement, _ in _ACCENT_RULES]


def fix_missing_accents(value: str) -> str:
    """Restore missing Spanish tildes/accents to words that were typed without them."""
    if not value:
        return value
    t = value
    for compiled_re, replacement in _COMPILED_ACCENT_RULES:
        t = compiled_re.sub(replacement, t)
    return t


@functools.lru_cache(maxsize=65536)
def _clean_visible_str(text: str) -> str:
    """Cached core logic for clean_visible_value (string inputs only)."""
    t = fix_corrupted_encoding(text)
    t = fix_missing_accents(t)
    t = unicodedata.normalize("NFC", t)
    t = "".join(char for char in t if unicodedata.category(char) not in {"Cc", "Cf"})
    t = re.sub(r"\s+", " ", t).strip()
    t = re.sub(r"\s*([,;:.])\s*", r"\1 ", t).strip()
    return t.rstrip(" ,;:.")


def clean_visible_value(value: Any) -> Any:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if not isinstance(value, str):
        return value
    return _clean_visible_str(value)


@functools.lru_cache(maxsize=65536)
def _comparison_str(text: str) -> str:
    """Cached core logic for comparison_value (string inputs only)."""
    t = unicodedata.normalize("NFKD", text)
    t = "".join(char for char in t if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]+", "", t.upper())


def comparison_value(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = float(value)
        return str(int(numeric)) if numeric.is_integer() else f"{numeric:.12g}"
    return _comparison_str(str(value))


def _read_csv(content: bytes) -> list[tuple[str, pd.DataFrame]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            frame = pd.read_csv(io.BytesIO(content), dtype=object, sep=None, engine="python", encoding=encoding)
            return [("CSV", frame)]
        except Exception as exc:  # pragma: no cover - only reached for malformed encodings
            last_error = exc
    raise CleaningError(f"No fue posible leer el archivo CSV: {last_error}")


def _read_excel(content: bytes, extension: str) -> list[tuple[str, pd.DataFrame]]:
    try:
        engine = "xlrd" if extension == ".xls" else "calamine"
        book = pd.ExcelFile(io.BytesIO(content), engine=engine)
        frames: list[tuple[str, pd.DataFrame]] = []
        for sheet in book.sheet_names:
            raw = pd.read_excel(book, sheet_name=sheet, header=None, dtype=object)
            frames.append((str(sheet), raw))
        return frames
    except Exception as exc:
        try:
            book = pd.ExcelFile(io.BytesIO(content), engine="openpyxl")
            frames = []
            for sheet in book.sheet_names:
                raw = pd.read_excel(book, sheet_name=sheet, header=None, dtype=object)
                frames.append((str(sheet), raw))
            return frames
        except Exception:
            raise CleaningError(f"No fue posible leer el archivo Excel: {exc}") from exc


def _candidate_frames(content: bytes, extension: str) -> list[tuple[str, int, pd.DataFrame]]:
    if extension == ".csv":
        return [("CSV", 0, frame) for _, frame in _read_csv(content)]

    candidates: list[tuple[str, int, pd.DataFrame]] = []
    for sheet, raw in _read_excel(content, extension):
        if raw.empty:
            continue
        best_header_idx = 0
        best_score = -1
        max_scan = min(25, len(raw.index))
        for idx in range(max_scan):
            hdr = [clean_visible_value(val) for val in raw.iloc[idx].tolist()]
            norm_set = {normalize_key(h) for h in hdr if normalize_key(h)}
            score = len(norm_set)
            if score > best_score:
                best_score = score
                best_header_idx = idx

        header = [clean_visible_value(value) for value in raw.iloc[best_header_idx].tolist()]
        frame = raw.iloc[best_header_idx + 1 :].copy()
        frame.columns = header
        candidates.append((sheet, best_header_idx, frame))
    return candidates


def _column_lookup(frame: pd.DataFrame) -> dict[str, Any]:
    lookup: dict[str, Any] = {}
    for column in frame.columns:
        key = normalize_key(column)
        if key and key not in lookup:
            lookup[key] = column
    return lookup


def _resolve_columns(frame: pd.DataFrame, required_headers: list[str]) -> dict[str, Any]:
    lookup = _column_lookup(frame)
    resolved: dict[str, Any] = {}
    for header in required_headers:
        aliases = [header, *HEADER_ALIASES.get(header, [])]
        source = next((lookup.get(normalize_key(alias)) for alias in aliases if lookup.get(normalize_key(alias)) is not None), None)
        if source is not None:
            resolved[header] = source
    return resolved


def _select_source(content: bytes, extension: str, required_headers: list[str]) -> tuple[str, pd.DataFrame, dict[str, Any]]:
    minimum = max(2, int(len(required_headers) * 0.20))
    candidate_frames = _candidate_frames(content, extension)
    matched_sources: list[tuple[str, pd.DataFrame, dict[str, Any]]] = []

    for sheet, _, frame in candidate_frames:
        resolved = _resolve_columns(frame, required_headers)
        if len(resolved) >= minimum:
            matched_sources.append((sheet, frame, resolved))

    if not matched_sources:
        best: tuple[int, str, pd.DataFrame, dict[str, Any]] | None = None
        for sheet, _, frame in candidate_frames:
            resolved = _resolve_columns(frame, required_headers)
            score = len(resolved)
            if best is None or score > best[0]:
                best = (score, sheet, frame, resolved)
        if not best or best[0] == 0:
            raise CleaningError("No se reconocieron columnas compatibles con la lista seleccionada.")
        missing = [header for header in required_headers if header not in best[3]][:8]
        raise CleaningError(
            f"La estructura no corresponde a la lista seleccionada. Solo se reconocieron {best[0]} de "
            f"{len(required_headers)} columnas. Faltan, entre otras: {', '.join(missing)}."
        )
    if len(matched_sources) == 1:
        return matched_sources[0][0], matched_sources[0][1], matched_sources[0][2]

    sheet_names = ", ".join(s[0] for s in matched_sources)
    combined_rows: list[pd.DataFrame] = []
    for sheet, frame, resolved in matched_sources:
        sub_df = pd.DataFrame({
            header: frame[resolved[header]] if header in resolved else ""
            for header in required_headers
        })
        combined_rows.append(sub_df)

    unified_frame = pd.concat(combined_rows, ignore_index=True)
    unified_resolved = {h: h for h in required_headers if any(h in r for _, _, r in matched_sources)}
    return sheet_names, unified_frame, unified_resolved


def _rules_index(rules: list[dict[str, Any]] | None) -> dict[tuple[str, str], str]:
    index: dict[tuple[str, str], str] = {}
    for rule in sorted(rules or [], key=lambda item: int(item.get("prioridad") or 100)):
        detected = comparison_value(rule.get("valor_detectado"))
        standard = clean_visible_value(rule.get("valor_estandar"))
        column = normalize_key(rule.get("columna") or "*")
        if detected and standard and (column, detected) not in index:
            index[(column, detected)] = standard
    return index


def _rule_scopes(header: str) -> list[str]:
    scopes = [normalize_key(header)]
    for _, value_header, scope in CANONICAL_RELATIONS:
        if value_header == header:
            scopes.append(normalize_key(scope))
    scopes.append("*")
    return list(dict.fromkeys(scopes))


def _apply_rule(header: str, value: Any, rules: dict[tuple[str, str], str]) -> Any:
    if not rules:
        return value
    detected = comparison_value(value)
    if not detected:
        return value
    for scope in _rule_scopes(header):
        standard = rules.get((scope, detected))
        if standard not in (None, ""):
            return standard
    return value


@functools.lru_cache(maxsize=1)
def _load_universities_reference_table() -> tuple[dict[tuple[str, str], dict[str, str]], dict[str, dict[str, str]]]:
    paths = [
        r"E:\Anderson\1. INDICADORES DE RENOVACIÓN\BASES DE DATOS\5. CONTEXTO EXTERNO\BASE GENERAL\Universidades.xlsx",
        os.path.join(os.path.dirname(__file__), "..", "data", "Universidades.xlsx")
    ]
    target_path = next((p for p in paths if os.path.exists(p)), None)
    pair_map: dict[tuple[str, str], dict[str, str]] = {}
    code_map: dict[str, dict[str, str]] = {}
    if not target_path:
        return pair_map, code_map

    try:
        df = pd.read_excel(target_path, dtype=str)
        cols_norm = {col: normalize_key(col) for col in df.columns}
        df = df.rename(columns=cols_norm)

        padre_col = next((c for c in df.columns if "PADRE" in c), None)
        inst_col = next((c for c in df.columns if ("INSTITUCION" in c or "CODIGO" in c) and "PADRE" not in c and "NOMBRE" not in c), None)
        nombre_col = next((c for c in df.columns if "NOMBRE" in c), None)
        caracter_col = next((c for c in df.columns if "CARACTER" in c), None)
        sector_col = next((c for c in df.columns if "SECTOR" in c), None)

        for _, row in df.iterrows():
            p_code = str(row.get(padre_col) or "").strip().split('.')[0] if padre_col else ""
            i_code = str(row.get(inst_col) or "").strip().split('.')[0] if inst_col else ""
            nombre = str(row.get(nombre_col) or "").strip().upper() if nombre_col else ""
            caracter = str(row.get(caracter_col) or "").strip().upper() if caracter_col else ""
            sector = str(row.get(sector_col) or "").strip().upper() if sector_col else ""

            if not nombre:
                continue

            entry = {"NOMBRE": nombre, "CARACTER": caracter, "SECTOR": sector}
            if p_code and i_code:
                pair_map[(p_code, i_code)] = entry
            if i_code:
                code_map[i_code] = entry
            if p_code and p_code not in code_map:
                code_map[p_code] = entry
    except Exception as exc:
        print(f"Error cargando tabla de referencia de Universidades: {exc}")

    return pair_map, code_map


@functools.lru_cache(maxsize=1)
def _load_snies_programs_reference_table() -> dict[str, dict[str, str]]:
    paths = [
        r"E:\Anderson\POWER BI\PLANEACIÓN Y EFECTIVIDAD\BASES DE DATOS\OFERTA_PROGRAMAS_SNIES.xlsx",
        os.path.join(os.path.dirname(__file__), "..", "data", "OFERTA_PROGRAMAS_SNIES.xlsx")
    ]
    target_path = next((p for p in paths if os.path.exists(p)), None)
    program_map: dict[str, dict[str, str]] = {}
    if not target_path:
        return program_map

    try:
        df = pd.read_excel(target_path, sheet_name="BD_OFERTA", dtype=str)
        for _, row in df.iterrows():
            snies_code = str(row.get("CODIGO_SNIES_DEL_PROGRAMA") or "").strip().split('.')[0]
            prog_name = str(row.get("NOMBRE_DEL_PROGRAMA") or "").strip().upper()
            if snies_code and prog_name and snies_code != "NAN":
                program_map[snies_code] = {
                    "PROGRAMA ACADEMICO": prog_name,
                    "NIVEL ACADEMICO": str(row.get("NIVEL_ACADEMICO") or "").strip().upper(),
                    "NIVEL DE FORMACION": str(row.get("NIVEL_DE_FORMACION") or "").strip().upper(),
                    "AREA DE CONOCIMIENTO": str(row.get("AREA_DE_CONOCIMIENTO") or "").strip().upper(),
                    "NUCLEO BASICO DEL CONOCIMIENTO (NBC)": str(row.get("NUCLEO_BASICO_DEL_CONOCIMIENTO") or "").strip().upper(),
                    "DESC CINE CAMPO AMPLIO": str(row.get("CINE_F_2013_AC_CAMPO_AMPLIO") or "").strip().upper(),
                    "DESC CINE CAMPO ESPECIFICO": str(row.get("CINE_F_2013_AC_CAMPO_ESPECIFIC") or "").strip().upper(),
                    "DESC CINE CAMPO DETALLADO": str(row.get("CINE_F_2013_AC_CAMPO_DETALLADO") or "").strip().upper(),
                    "DEPARTAMENTO DE OFERTA DEL PROGRAMA": str(row.get("DEPARTAMENTO_OFERTA_PROGRAMA") or "").strip().upper(),
                    "MUNICIPIO DE OFERTA DEL PROGRAMA": str(row.get("MUNICIPIO_OFERTA_PROGRAMA") or "").strip().upper(),
                }
    except Exception as exc:
        print(f"Error cargando tabla de oferta SNIES: {exc}")

    return program_map


@functools.lru_cache(maxsize=1)
def _load_divipola_reference_table() -> tuple[dict[str, str], dict[str, str]]:
    paths = [
        r"E:\Anderson\3. SNIES\CODÍGOS DIVIPOLA.xlsx",
        os.path.join(os.path.dirname(__file__), "..", "data", "CODÍGOS_DIVIPOLA.xlsx")
    ]
    target_path = next((p for p in paths if os.path.exists(p)), None)
    dept_map: dict[str, str] = {}
    muni_map: dict[str, str] = {}
    if not target_path:
        return dept_map, muni_map

    try:
        df = pd.read_excel(target_path, sheet_name=0, header=7, dtype=str)
        for _, row in df.iterrows():
            d_code = str(row.iloc[0] or "").strip().split('.')[0]
            d_name = str(row.iloc[2] or "").strip().upper()
            m_code = str(row.iloc[1] or "").strip().split('.')[0]
            m_name = str(row.iloc[3] or "").strip().upper()

            if d_code and d_name and d_code != "NAN":
                dept_map[d_code] = d_name
                dept_map[d_code.zfill(2)] = d_name
            if m_code and m_name and m_code != "NAN":
                muni_map[m_code] = m_name
                muni_map[m_code.zfill(5)] = m_name
    except Exception as exc:
        print(f"Error cargando tabla DIVIPOLA DANE: {exc}")

    return dept_map, muni_map


def _quality_score(value: str) -> tuple[int, int, int, int]:
    text = str(value or "")
    letters = sum(char.isalpha() for char in text)
    suspicious = len(re.findall(r"(.)\1{2,}", text, flags=re.IGNORECASE))
    accents = sum(unicodedata.normalize("NFD", char) != char for char in text)
    return (-suspicious, accents, letters, -len(text))


def _word_tokens(value: str) -> list[str]:
    return re.findall(r"[^\W_]+", unicodedata.normalize("NFC", str(value or "")), flags=re.UNICODE)


def _has_diacritic(value: str) -> bool:
    return any(unicodedata.combining(char) for char in unicodedata.normalize("NFD", value))


def _accent_consensus(counter: Counter) -> dict[str, str]:
    candidates: dict[str, Counter] = defaultdict(Counter)
    for value, count in counter.items():
        for token in _word_tokens(value):
            if _has_diacritic(token):
                candidates[comparison_value(token)][unicodedata.normalize("NFC", token).upper()] += count
    return {
        key: max(options, key=lambda token: (options[token], -len(token), token))
        for key, options in candidates.items()
    }


def _accent_agreement(value: str, consensus: dict[str, str]) -> tuple[int, int]:
    matches = 0
    misplaced = 0
    for token in _word_tokens(value):
        preferred = consensus.get(comparison_value(token))
        if not preferred:
            continue
        normalized = unicodedata.normalize("NFC", token).upper()
        if normalized == preferred:
            matches += 1
        elif _has_diacritic(token):
            misplaced += 1
    return matches, -misplaced


def _choose_canonical(counter: Counter) -> str:
    semantic_counts: Counter = Counter()
    variants: dict[str, Counter] = defaultdict(Counter)
    for value, count in counter.items():
        key = comparison_value(value)
        if not key:
            continue
        semantic_counts[key] += count
        variants[key][value] += count
    if not semantic_counts:
        return ""
    best_key = max(semantic_counts, key=lambda key: (semantic_counts[key], max(_quality_score(v) for v in variants[key]), len(key)))
    accent_consensus = _accent_consensus(variants[best_key])
    return max(
        variants[best_key],
        key=lambda value: (
            _accent_agreement(value, accent_consensus),
            variants[best_key][value],
            _quality_score(value)[0],
            _quality_score(value)[2:],
        ),
    )


def _excel_source(value: bytes | BinaryIO) -> BinaryIO:
    if isinstance(value, (bytes, bytearray)):
        return io.BytesIO(value)
    value.seek(0)
    return value


def _find_excel_source(workbook, required_headers: list[str]):
    best = None
    for worksheet in workbook.worksheets:
        for header_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
            lookup = {normalize_key(value): index for index, value in enumerate(row) if normalize_key(value)}
            resolved = {}
            for header in required_headers:
                aliases = [header, *HEADER_ALIASES.get(header, [])]
                source_index = next((lookup.get(normalize_key(alias)) for alias in aliases if normalize_key(alias) in lookup), None)
                if source_index is not None:
                    resolved[header] = source_index
            candidate = (len(resolved), worksheet, header_index, resolved)
            if best is None or candidate[0] > best[0]:
                best = candidate
    minimum = max(2, int(len(required_headers) * 0.35))
    if not best or best[0] < minimum:
        raise CleaningError("No se encontró una hoja con la estructura correspondiente a la lista seleccionada.")
    return best[1], best[2], best[3]


def _iter_excel_rows(worksheet, header_row: int, resolved: dict[str, int], headers: list[str], rules):
    for source_row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        values = []
        originals = []
        for header in headers:
            index = resolved.get(header)
            original = source_row[index] if index is not None and index < len(source_row) else ""
            visible = clean_visible_value(original)
            values.append(_apply_rule(header, visible, rules))
            originals.append(original)
        yield originals, values


def _clean_excel_streaming(
    content: bytes | BinaryIO,
    required_headers: list[str],
    list_key: str,
    rules: list[dict[str, Any]] | None,
    output_path: str | None = None,
) -> CleaningResult:
    if not isinstance(content, (bytes, bytearray)):
        content.seek(0)
        content = content.read()

    sheet, source, resolved = _select_source(content, ".xlsx", required_headers)
    rule_index = _rules_index(rules)

    raw_df = pd.DataFrame({
        header: source[resolved[header]] if header in resolved else ""
        for header in required_headers
    })

    input_rows = len(raw_df)
    cleaned_data = {}
    original_data = {}
    for header in required_headers:
        orig_cleaned = raw_df[header].map(clean_visible_value)
        original_data[header] = orig_cleaned
        cleaned_data[header] = orig_cleaned.map(lambda val: _apply_rule(header, val, rule_index))

    df = pd.DataFrame(cleaned_data)
    orig_df = pd.DataFrame(original_data)

    nonempty_mask = (df != "").any(axis=1)
    empty_rows = int((~nonempty_mask).sum())
    df = df.loc[nonempty_mask].copy().reset_index(drop=True)
    orig_df = orig_df.loc[nonempty_mask].copy().reset_index(drop=True)
    output_rows = len(df)

    relation_candidates: dict[tuple[str, str, str], Counter] = defaultdict(Counter)
    for code_header, value_header, _ in CANONICAL_RELATIONS:
        if code_header in df.columns and value_header in df.columns:
            valid_sub = df[[code_header, value_header]].dropna()
            if not valid_sub.empty:
                counts = valid_sub.groupby([code_header, value_header], observed=True).size()
                for (c_val, v_val), count in counts.items():
                    code = comparison_value(c_val)
                    val = clean_visible_value(v_val)
                    if code and comparison_value(val):
                        relation_candidates[(code_header, value_header, code)][val] += int(count)
    canonical = {key: _choose_canonical(counter) for key, counter in relation_candidates.items()}

    global_variant_candidates: dict[tuple[str, str], Counter] = defaultdict(Counter)
    for header in GLOBAL_CANONICAL_HEADERS.intersection(required_headers):
        if header in df.columns:
            series_counts = df[header].dropna().value_counts()
            for val, count in series_counts.items():
                val_clean = clean_visible_value(val)
                sem_key = comparison_value(val_clean)
                if sem_key:
                    global_variant_candidates[(header, sem_key)][val_clean] += int(count)
    global_canonical = {key: _choose_canonical(counter) for key, counter in global_variant_candidates.items()}

    corrections: Counter = Counter()
    explicitly_corrected: dict[str, set[int]] = {header: set() for header in required_headers}

    # === TABLA DE REFERENCIA DE UNIVERSIDADES (DOBLE VALIDACIÓN POR CÓDIGOS) ===
    pair_ref_map, code_ref_map = _load_universities_reference_table()
    if pair_ref_map or code_ref_map:
        p_header = "CÓDIGO_INSTITUCIÓN_PADRE" if "CÓDIGO_INSTITUCIÓN_PADRE" in df.columns else ("IES PADRE" if "IES PADRE" in df.columns else None)
        i_header = "CÓDIGO_INSTITUCIÓN" if "CÓDIGO_INSTITUCIÓN" in df.columns else ("CODIGO DE LA INSTITUCION" if "CODIGO DE LA INSTITUCION" in df.columns else None)
        name_header = "Institución de Educación Superior (IES)" if "Institución de Educación Superior (IES)" in df.columns else ("INSTITUCION DE EDUCACION SUPERIOR (IES)" if "INSTITUCION DE EDUCACION SUPERIOR (IES)" in df.columns else None)

        if name_header and (p_header or i_header):
            p_s = df[p_header].astype(str).str.split('.').str[0].str.strip() if p_header else pd.Series([""] * len(df))
            i_s = df[i_header].astype(str).str.split('.').str[0].str.strip() if i_header else pd.Series([""] * len(df))

            keys = list(zip(p_s, i_s))
            matched_entries = [
                pair_ref_map.get(k) or code_ref_map.get(k[1]) or code_ref_map.get(k[0])
                for k in keys
            ]

            # Standardize Institution Name
            names_std = pd.Series([e["NOMBRE"] if e else None for e in matched_entries], index=df.index)
            mask_name = names_std.notna() & (names_std.astype(str) != df[name_header].astype(str))
            if mask_name.any():
                diff_df = pd.DataFrame({'code': p_s.loc[mask_name] + "/" + i_s.loc[mask_name], 'cur': df.loc[mask_name, name_header], 'std': names_std.loc[mask_name]})
                for (code, cur, std), count in diff_df.groupby(['code', 'cur', 'std'], observed=True).size().items():
                    corrections[("IES", str(code), str(cur).upper(), str(std).upper(), "TABLA_REFERENCIA_UNIVERSIDADES")] += int(count)
                df.loc[mask_name, name_header] = names_std.loc[mask_name]
                for idx in df.index[mask_name]:
                    explicitly_corrected[name_header].add(idx)

            # Standardize Sector IES
            sec_hdr = "Sector IES" if "Sector IES" in df.columns else ("SECTOR IES" if "SECTOR IES" in df.columns else None)
            if sec_hdr:
                sector_std = pd.Series([e["SECTOR"] if e and e.get("SECTOR") else None for e in matched_entries], index=df.index)
                mask_sec = sector_std.notna() & (sector_std.astype(str) != df[sec_hdr].astype(str))
                if mask_sec.any():
                    diff_df = pd.DataFrame({'code': p_s.loc[mask_sec] + "/" + i_s.loc[mask_sec], 'cur': df.loc[mask_sec, sec_hdr], 'std': sector_std.loc[mask_sec]})
                    for (code, cur, std), count in diff_df.groupby(['code', 'cur', 'std'], observed=True).size().items():
                        corrections[("SECTOR_IES", str(code), str(cur).upper(), str(std).upper(), "TABLA_REFERENCIA_UNIVERSIDADES")] += int(count)
                    df.loc[mask_sec, sec_hdr] = sector_std.loc[mask_sec]
                    for idx in df.index[mask_sec]:
                        explicitly_corrected[sec_hdr].add(idx)

            # Standardize Caracter IES
            car_hdr = "Caracter IES" if "Caracter IES" in df.columns else ("CARACTER IES" if "CARACTER IES" in df.columns else None)
            if car_hdr:
                caracter_std = pd.Series([e["CARACTER"] if e and e.get("CARACTER") else None for e in matched_entries], index=df.index)
                mask_car = caracter_std.notna() & (caracter_std.astype(str) != df[car_hdr].astype(str))
                if mask_car.any():
                    diff_df = pd.DataFrame({'code': p_s.loc[mask_car] + "/" + i_s.loc[mask_car], 'cur': df.loc[mask_car, car_hdr], 'std': caracter_std.loc[mask_car]})
                    for (code, cur, std), count in diff_df.groupby(['code', 'cur', 'std'], observed=True).size().items():
                        corrections[("CARACTER_IES", str(code), str(cur).upper(), str(std).upper(), "TABLA_REFERENCIA_UNIVERSIDADES")] += int(count)
                    df.loc[mask_car, car_hdr] = caracter_std.loc[mask_car]
                    for idx in df.index[mask_car]:
                        explicitly_corrected[car_hdr].add(idx)

    # === TABLA DE REFERENCIA DE METADATOS OFICIALES SNIES (CÓDIGOS E IDs) ===
    for (id_hdr, desc_hdr), meta_dict in OFFICIAL_METADATA_MAPS.items():
        if id_hdr in df.columns and desc_hdr in df.columns:
            id_series = df[id_hdr].astype(str).str.split('.').str[0].str.strip()
            std_desc = id_series.map(meta_dict)
            mask_meta = std_desc.notna() & (std_desc.astype(str) != df[desc_hdr].astype(str))
            if mask_meta.any():
                diff_df = pd.DataFrame({'code': id_series.loc[mask_meta], 'cur': df.loc[mask_meta, desc_hdr], 'std': std_desc.loc[mask_meta]})
                for (code, cur, std), count in diff_df.groupby(['code', 'cur', 'std'], observed=True).size().items():
                    corrections[(normalize_key(desc_hdr), str(code), str(cur).upper(), str(std).upper(), "TABLA_REFERENCIA_METADATOS")] += int(count)
                df.loc[mask_meta, desc_hdr] = std_desc.loc[mask_meta]
                for idx in df.index[mask_meta]:
                    explicitly_corrected[desc_hdr].add(idx)

    # === TABLA DE REFERENCIA DE PROGRAMAS SNIES (30,445 PROGRAMAS OFICIALES) ===
    prog_ref_map = _load_snies_programs_reference_table()
    snies_hdr = "Código SNIES del programa" if "Código SNIES del programa" in df.columns else ("CODIGO SNIES DEL PROGRAMA" if "CODIGO SNIES DEL PROGRAMA" in df.columns else None)
    if prog_ref_map and snies_hdr:
        snies_s = df[snies_hdr].astype(str).str.split('.').str[0].str.strip()
        matched_progs = [prog_ref_map.get(code) for code in snies_s]

        prog_cols_to_check = [
            ("Programa Académico", "PROGRAMA ACADEMICO"),
            ("Nivel Académico", "NIVEL ACADEMICO"),
            ("Nivel de Formación", "NIVEL DE FORMACION"),
            ("Área de Conocimiento", "AREA DE CONOCIMIENTO"),
            ("Núcleo Básico del Conocimiento (NBC)", "NUCLEO BASICO DEL CONOCIMIENTO (NBC)"),
            ("DESC CINE CAMPO AMPLIO", "DESC CINE CAMPO AMPLIO"),
            ("DESC CINE CAMPO ESPECIFICO", "DESC CINE CAMPO ESPECIFICO"),
            ("DESC CINE CODIGO DETALLADO", "DESC CINE CODIGO DETALLADO"),
            ("Departamento de oferta del programa", "DEPARTAMENTO DE OFERTA DEL PROGRAMA"),
            ("Municipio de oferta del programa", "MUNICIPIO DE OFERTA DEL PROGRAMA"),
        ]

        for df_col, map_key in prog_cols_to_check:
            if df_col in df.columns:
                std_series = pd.Series([p[map_key] if p and p.get(map_key) else None for p in matched_progs], index=df.index)
                mask_p = std_series.notna() & (std_series.astype(str) != df[df_col].astype(str))
                if mask_p.any():
                    diff_df = pd.DataFrame({'code': snies_s.loc[mask_p], 'cur': df.loc[mask_p, df_col], 'std': std_series.loc[mask_p]})
                    for (code, cur, std), count in diff_df.groupby(['code', 'cur', 'std'], observed=True).size().items():
                        corrections[(normalize_key(df_col), str(code), str(cur).upper(), str(std).upper(), "TABLA_REFERENCIA_PROGRAMAS_SNIES")] += int(count)
                    df.loc[mask_p, df_col] = std_series.loc[mask_p]
                    for idx in df.index[mask_p]:
                        explicitly_corrected[df_col].add(idx)

    # === TABLA DE REFERENCIA DIVIPOLA DANE (1,269 MUNICIPIOS Y 35 DEPARTAMENTOS) ===
    dept_divipola, muni_divipola = _load_divipola_reference_table()
    if dept_divipola or muni_divipola:
        dept_cols = [
            ("Código del departamento (IES)", "Departamento de domicilio de la IES"),
            ("Código del Departamento (Programa)", "Departamento de oferta del programa")
        ]
        for code_col, name_col in dept_cols:
            if code_col in df.columns and name_col in df.columns:
                codes_s = df[code_col].astype(str).str.split('.').str[0].str.strip()
                std_names = codes_s.map(dept_divipola)
                mask_d = std_names.notna() & (std_names.astype(str) != df[name_col].astype(str))
                if mask_d.any():
                    diff_df = pd.DataFrame({'code': codes_s.loc[mask_d], 'cur': df.loc[mask_d, name_col], 'std': std_names.loc[mask_d]})
                    for (code, cur, std), count in diff_df.groupby(['code', 'cur', 'std'], observed=True).size().items():
                        corrections[(normalize_key(name_col), str(code), str(cur).upper(), str(std).upper(), "TABLA_REFERENCIA_DIVIPOLA_DANE")] += int(count)
                    df.loc[mask_d, name_col] = std_names.loc[mask_d]
                    for idx in df.index[mask_d]:
                        explicitly_corrected[name_col].add(idx)

        muni_cols = [
            ("Código del Municipio (IES)", "Municipio de domicilio de la IES"),
            ("Código del Municipio (Programa)", "Municipio de oferta del programa")
        ]
        for code_col, name_col in muni_cols:
            if code_col in df.columns and name_col in df.columns:
                codes_s = df[code_col].astype(str).str.split('.').str[0].str.strip()
                std_names = codes_s.map(muni_divipola)
                mask_m = std_names.notna() & (std_names.astype(str) != df[name_col].astype(str))
                if mask_m.any():
                    diff_df = pd.DataFrame({'code': codes_s.loc[mask_m], 'cur': df.loc[mask_m, name_col], 'std': std_names.loc[mask_m]})
                    for (code, cur, std), count in diff_df.groupby(['code', 'cur', 'std'], observed=True).size().items():
                        corrections[(normalize_key(name_col), str(code), str(cur).upper(), str(std).upper(), "TABLA_REFERENCIA_DIVIPOLA_DANE")] += int(count)
                    df.loc[mask_m, name_col] = std_names.loc[mask_m]
                    for idx in df.index[mask_m]:
                        explicitly_corrected[name_col].add(idx)

    for code_header, value_header, scope in CANONICAL_RELATIONS:
        if code_header in df.columns and value_header in df.columns:
            mapping = {code: std for (ch, vh, code), std in canonical.items() if ch == code_header and vh == value_header and std}
            if mapping:
                code_series = df[code_header].map(comparison_value)
                new_vals = code_series.map(mapping)
                mask = new_vals.notna() & (new_vals.astype(str) != df[value_header].astype(str))
                if mask.any():
                    diff_df = pd.DataFrame({
                        'code': code_series.loc[mask],
                        'cur': df.loc[mask, value_header],
                        'std': new_vals.loc[mask]
                    })
                    pair_counts = diff_df.groupby(['code', 'cur', 'std'], observed=True).size()
                    for (code, cur, std), count in pair_counts.items():
                        corrections[(scope, str(code), str(cur).upper(), str(std).upper(), "CANONIZACION_POR_CODIGO")] += int(count)
                    df.loc[mask, value_header] = new_vals.loc[mask]
                    for idx in df.index[mask]:
                        explicitly_corrected[value_header].add(idx)

    for header in GLOBAL_CANONICAL_HEADERS.intersection(required_headers):
        if header in df.columns:
            mapping = {sem_key: std for (h, sem_key), std in global_canonical.items() if h == header and std}
            if mapping:
                val_series = df[header].map(clean_visible_value)
                sem_series = val_series.map(comparison_value)
                new_vals = sem_series.map(mapping)
                mask = new_vals.notna() & (new_vals.astype(str) != df[header].astype(str))
                if mask.any():
                    diff_df = pd.DataFrame({
                        'cur': val_series.loc[mask],
                        'std': new_vals.loc[mask]
                    })
                    pair_counts = diff_df.groupby(['cur', 'std'], observed=True).size()
                    for (cur, std), count in pair_counts.items():
                        corrections[(normalize_key(header), "", str(cur).upper(), str(std).upper(), "NORMALIZACION_ORTOGRAFICA")] += int(count)
                    df.loc[mask, header] = new_vals.loc[mask]
                    for idx in df.index[mask]:
                        explicitly_corrected[header].add(idx)

    for header in required_headers:
        raw_series = raw_df.loc[nonempty_mask, header].reset_index(drop=True)
        cleaned_series = orig_df[header]
        mojibake_mask = raw_series.astype(str).str.contains(r'[¿?\ufffd]', regex=True, na=False) & ~raw_series.astype(str).str.contains(r'^\s*\?\s*$', regex=True, na=False)
        if mojibake_mask.any():
            diff_df = pd.DataFrame({'raw': raw_series.loc[mojibake_mask], 'repaired': cleaned_series.loc[mojibake_mask]})
            diff_df = diff_df[diff_df['raw'].astype(str).str.strip() != diff_df['repaired'].astype(str).str.strip()]
            if not diff_df.empty:
                pair_counts = diff_df.groupby(['raw', 'repaired'], observed=True).size()
                for (raw_val, repaired_val), count in pair_counts.items():
                    corrections[(normalize_key(header), "", str(raw_val).upper(), str(repaired_val).upper(), "CORRECCION_MOJIBAKE")] += int(count)

    for header in required_headers:
        orig_series = orig_df[header].map(clean_visible_value)
        final_series = df[header]
        mask = (orig_series.map(comparison_value) != "") & (orig_series.astype(str) != final_series.astype(str))
        if mask.any():
            diff_df = pd.DataFrame({'orig': orig_series.loc[mask], 'final': final_series.loc[mask]})
            pair_counts = diff_df.groupby(['orig', 'final'], observed=True).size()
            for (orig, final), count in pair_counts.items():
                reason = "NORMALIZACION_TECNICA" if comparison_value(orig) == comparison_value(final) else "REGLA_DICCIONARIO"
                corrections[(normalize_key(header), "", str(orig).upper(), str(final).upper(), reason)] += int(count)

    for header in required_headers:
        if header not in NUMERIC_HEADERS:
            df[header] = df[header].map(lambda v: str(v).upper() if (v is not None and not (isinstance(v, float) and pd.isna(v))) else "")

    for header in NUMERIC_HEADERS.intersection(df.columns):
        df[header] = pd.to_numeric(df[header], errors="coerce").where(lambda s: s.notna(), "")

    output_buffer = io.BytesIO() if not output_path else None
    sheet_name = list_key[:31].upper()

    correction_rows = [
        {
            "COLUMNA": column.upper(),
            "CODIGO_REFERENCIA": code.upper() if isinstance(code, str) else code,
            "VALOR_DETECTADO": detected.upper() if isinstance(detected, str) else detected,
            "VALOR_ESTANDAR": standard.upper() if isinstance(standard, str) else standard,
            "MOTIVO": reason,
            "OCURRENCIAS": count,
        }
        for (column, code, detected, standard, reason), count in corrections.most_common()
    ]
    report_df = pd.DataFrame(correction_rows if correction_rows else [], columns=["COLUMNA", "CODIGO_REFERENCIA", "VALOR_DETECTADO", "VALOR_ESTANDAR", "MOTIVO", "OCURRENCIAS"])

    write_target = output_path if output_path else output_buffer
    with pd.ExcelWriter(write_target, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        report_df.to_excel(writer, sheet_name="REPORTE_NORMALIZACION", index=False)

    clean_bytes = b"" if output_path else output_buffer.getvalue()

    return CleaningResult(
        content=clean_bytes,
        input_rows=input_rows,
        output_rows=output_rows,
        duplicates_removed=0,
        empty_rows_removed=empty_rows,
        source_sheet=sheet,
        matched_columns=len(resolved),
        corrections_count=sum(corrections.values()),
        corrections=correction_rows,
    )


def clean_contexto_externo_file(
    content: bytes | BinaryIO,
    filename: str,
    list_name: str,
    rules: list[dict[str, Any]] | None = None,
    output_path: str | None = None,
) -> CleaningResult:
    list_key = normalize_key(list_name)
    required_headers = LIST_CONFIG.get(list_key)
    if not required_headers:
        raise CleaningError("La lista de Contexto Externo seleccionada no es válida.")

    extension_match = re.search(r"(\.xlsx|\.xls|\.csv)$", str(filename or ""), flags=re.IGNORECASE)
    if not extension_match:
        raise CleaningError("Solo se admiten archivos .xlsx, .xls o .csv.")
    extension = extension_match.group(1).lower()

    if extension == ".xlsx":
        return _clean_excel_streaming(content, required_headers, list_key, rules, output_path=output_path)

    if not isinstance(content, (bytes, bytearray)):
        content.seek(0)
        content = content.read()

    sheet, source, resolved = _select_source(content, extension, required_headers)
    output = pd.DataFrame({
        header: source[resolved[header]] if header in resolved else ""
        for header in required_headers
    })
    input_rows = len(output.index)
    output = output.map(clean_visible_value)

    nonempty_mask = output.apply(lambda row: any(comparison_value(value) for value in row), axis=1)
    empty_rows_removed = int((~nonempty_mask).sum())
    output = output.loc[nonempty_mask].copy()

    for header in NUMERIC_HEADERS.intersection(output.columns):
        output[header] = pd.to_numeric(output[header], errors="coerce").where(lambda series: series.notna(), "")

    output = output.reset_index(drop=True)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        output.to_excel(writer, index=False, sheet_name=list_key[:31])
        worksheet = writer.book[list_key[:31]]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="2563EB")
        for column_cells in worksheet.columns:
            width = min(45, max(12, max(len(str(cell.value or "")) for cell in list(column_cells)[:200]) + 2))
            worksheet.column_dimensions[column_cells[0].column_letter].width = width

    return CleaningResult(
        content=buffer.getvalue(),
        input_rows=input_rows,
        output_rows=len(output.index),
        duplicates_removed=0,
        empty_rows_removed=empty_rows_removed,
        source_sheet=sheet,
        matched_columns=len(resolved),
        corrections_count=0,
        corrections=[],
    )
